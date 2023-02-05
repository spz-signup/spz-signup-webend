# -*- coding: utf-8 -*-

"""Formatter that writes excel files.
"""

from . import TableWriter

import re

from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl.worksheet.table import Table
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from zipfile import ZipFile
from spz import app


def find_table(workbook, table_name):
    for sheet in workbook.worksheets:
        for table in sheet._tables:
            return sheet, table

    """function to find information table with the course information
    in table: Level, ECTS, course name, date of exam
    """


def find_course_table(name, tables):
    for table in tables:
        if table.displayName == name:
            return table.ref



def delete_last_row(sheet, range):
    for c in range.bottom:
        sheet.cell(*c).value = None
    range.shrink(bottom=1)


def sanitize_title(str):
    """Sanitize a string so it can safely be used as a worksheet title:
    Disallowed characters will be stripped and the string will be trimmed to a length of 30.
    """
    return re.sub(INVALID_TITLE_REGEX, '', str)[:30]


class ExcelWriter(TableWriter):
    """The base ExcelWriter begins a new sheet for each new section.
    """

    @property
    def mimetype(self):
        return self.workbook.mime_type

    @property
    def extension(self):
        return 'xltx' if self.workbook.template else 'xlsx'

    def __init__(self, template):
        TableWriter.__init__(self, template, binary_template=True)
        self.section_count = 0

    def parse_template(self, file):
        self.workbook = load_workbook(file)
        # once we have the template sheet and its index, we remove it from the workbook
        # this way we can add multiple copies of it while keeping the original unmodified
        self.template_sheet, self.template_table = find_table(self.workbook, 'DATA')
        self.sheet_insert_index = self.workbook.index(self.template_sheet)
        self.template_range = CellRange(self.template_table.ref)
        expression_row = [self.template_sheet.cell(*c).value for c in self.template_range.bottom]
        delete_last_row(self.template_sheet, self.template_range)  # this row contains the jinja-expressions
        self.workbook.remove(self.template_sheet)
        return super().parse_template(expression_row)

    def set_course_information(self, course):
        pass

    def begin_section(self, section_name):
        title = sanitize_title(section_name)
        self.current_sheet = self.workbook.create_sheet(title=title, index=self.sheet_insert_index + self.section_count)
        self.current_range = CellRange(self.template_range.coord)
        self.section_count += 1  # insert sheets in ascending order
        WorksheetCopy(source_worksheet=self.template_sheet, target_worksheet=self.current_sheet).copy_worksheet()

    def end_section(self, section_name=None):
        # TODO: copy more properties of template_table
        # TODO: make sure that tables have at least one entry
        table = Table(ref=self.current_range.coord, displayName="DATA_{}".format(self.section_count))
        self.current_sheet.add_table(table)

    def write_row(self, row):
        row_iter = iter(row)
        self.current_range.expand(down=1)
        for c in self.current_range.bottom:
            self.current_sheet.cell(*c).value = next(row_iter)

    def get_data(self):
        with NamedTemporaryFile() as file:
            self.workbook.save(file.name)
            file.seek(0)
            stream = file.read()
        return stream


class ExcelZipWriter(ExcelWriter):
    """ The ExcelZipWriter begins a new .xlsx file for each new section.
    """

    mimetype = 'application/zip'
    extension = 'zip'

    def __init__(self, template):
        ExcelWriter.__init__(self, template)
        self.tempfile = NamedTemporaryFile()
        self.zip = ZipFile(self.tempfile, 'w')

    def set_course_information(self, course):
        # set course information
        self.current_sheet = self.workbook.get_sheet_by_name("Notenliste")
        course_information = []
        expressions = []
        starting_cell = None
        # iterate sheet to find jinja expressions
        for row in self.current_sheet.iter_rows(min_row=30, min_col=1, max_row=50, max_col=3):
            for cell in row:
                if cell.value is not None:
                    key = cell.value
                    options = [
                        'course.ger',
                        'course.ects_points',
                        'course.full_name',
                        'course.level',
                        'course.full_name_english',
                        'semester',
                        'exam_date'
                    ]
                    # if one of the strings is equal, it gets added to the information list
                    if any(key in word for word in options):
                        if len(expressions) == 0:
                            starting_cell = cell.coordinate
                        expressions.append(key)
                        cell.value = None
        #TODO: set semester and semester_date in json file globally and import this data here
        semester = "WS 2022-2023"
        exam_date = "17.02.2023"
        # gets converted into callable expression
        course_information = [app.jinja_env.compile_expression(e) for e in expressions]
        # convert jinja expressions into writable expression with the required data
        expression_column = [cell_template(dict(course=course, semester=semester, exam_date=exam_date))
                             for cell_template in course_information]

        # write the information column starting at the first found cell
        offset_count = 0
        for expression in expression_column:
            cell = self.current_sheet[starting_cell].offset(offset_count, 0)
            cell.value = expression
            offset_count += 1

        #TODO: calculate ects points (define globally in course)
        ects_points = None
        if course.price <= 90:
            ects_points = 2
        else:
            ects_points = 4


    def begin_section(self, section_name):
        # use title of template sheet
        super().begin_section(section_name=self.template_sheet.title)

    def end_section(self, section_name):
        super().end_section(section_name)
        with NamedTemporaryFile() as file:
            self.workbook.save(file.name)
            self.zip.write(file.name, "{}.xlsx".format(section_name))
        # Restore template workbook to initial state
        self.workbook.remove(self.current_sheet)
        self.section_count -= 1

    def get_data(self):
        self.zip.close()
        self.tempfile.seek(0)
        return self.tempfile.read()


class SingleSectionExcelWriter(ExcelWriter):
    section = None

    def set_course_information(self, course):
        pass

    def begin_section(self, section_name):
        if not self.section:
            super().begin_section(section_name)
        self.section = section_name

    def end_section(self, section_name):
        pass

    def get_data(self):
        if self.section:
            super().end_section(self.section)
        return super().get_data()
