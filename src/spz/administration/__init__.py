# -*- coding: utf-8 -*-
"""static functions related to administration tasks

This module contains methods for:
    - teacher management
    - course management
"""

from spz import app
from spz import models, db

from flask_babel import gettext as _
from flask import flash
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
import re


def get_course_ids():
    course_ids_tuple = db.session.query(models.Role.course_id).filter(
        models.Role.course_id.isnot(None)) \
        .filter(models.Role.role == models.Role.COURSE_TEACHER) \
        .distinct().all()
    # transform query tuples into right integer format
    """flash('Vergebene Kurse sind:')
    for course_id in course_ids_tuple:
        course = models.Course.query.get(course_id[0])
        flash(f'{course.full_name}: {course_id[0]}')"""
    return [course_id[0] for course_id in course_ids_tuple]


def to_float(value):
    try:
        if type(value) == float:
            return value
        elif type(value) == int:
            return float(value)
        elif type(value) == str:
            return float(value.strip().replace(',', '.'))
    except ValueError:
        return None


def is_valid_float(value):
    """Check if a value can be converted to a float."""
    try:
        # Attempt to convert the value to a float
        if isinstance(value, float):
            return True
        elif isinstance(value, int):
            return True
        elif isinstance(value, str):
            # Strip whitespace and replace commas with dots for string conversion
            float(value.strip().replace(',', '.'))
            return True
    except ValueError:
        pass
    # Return False if conversion fails
    return False


def validate_email(email):
    """Validate email format, similar to Flask-WTForms."""

    email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if re.match(email_regex, email) is None:
        return False

    return True


class TeacherManagement:
    @staticmethod
    def remove_course(teacher, course, user_id):
        role_to_remove = models.Role.query. \
            join(models.User.roles). \
            filter(models.Role.user_id == user_id). \
            filter(models.Role.course_id == course.id). \
            first()
        if role_to_remove:
            # remove complete table row from role table
            db.session.delete(role_to_remove)
        else:
            raise ValueError(_('Folgender Kurs "{}" war kein Kurs des/der Lehrbeauftragten.'
                               ' Wurde der richtige Kurs ausgewählt?'.format(course.full_name)))

        return course

    @staticmethod
    def add_course(teacher, course):
        own_courses_id = [course.id for course in teacher.teacher_courses]
        if course.id in own_courses_id:
            raise ValueError(
                _('Der/die Lehrbeauftragte hat diesen Kurs schon zugewiesen. Doppelzuweisung nicht möglich!'))
        TeacherManagement.check_availability(course)
        teacher.roles.append(models.Role(course=course, role=models.Role.COURSE_TEACHER))

    @staticmethod
    def check_availability(course):
        teachers = db.session.query(models.User) \
            .join(models.Role, models.User.roles) \
            .filter(models.Role.role == models.Role.COURSE_TEACHER).all()
        course_ids = get_course_ids()
        if course.id in course_ids:
            for teacher in teachers:
                if teacher.is_course_teacher(course):
                    # flash(f'{course.full_name} ({course.id}) is already assigned to {teacher.full_name}')
                    raise ValueError('{0} ist schon vergeben an {1}.'.format(course.full_name, teacher.full_name))

    @staticmethod
    def unassigned_courses(language_id):
        # courses with assigned teachers only
        unassigned_courses = db.session.query(models.Course) \
            .outerjoin(models.Role,
                       (models.Role.course_id == models.Course.id) & (models.Role.role == models.Role.COURSE_TEACHER)) \
            .join(models.Language) \
            .filter(models.Language.id == language_id) \
            .filter(models.Role.id == None)

        return unassigned_courses

    @staticmethod
    def import_grades(xlsx_file, course):

        grade_wb = load_workbook(xlsx_file, read_only=True, data_only=True)

        # check for 'Notenliste' and 'RAWDATA' sheet
        if not all(sheet in grade_wb.sheetnames for sheet in ['RAWDATA', 'Notenliste']):
            raise ValueError(
                _('Die Excel-Datei enthält nicht die notwendigen Blätter "RAWDATA" und/oder "Notenliste".'))

        # check course name in 'Notenliste' sheet -> throw warning if not matching with course
        grade_sheet = grade_wb['Notenliste']
        for row in grade_sheet.iter_rows(min_row=40, max_col=3, max_row=50):
            for cell in row:
                if cell.value == 'Kursname:':
                    read_course_name = grade_sheet.cell(row=cell.row, column=cell.column + 1).value
                    if read_course_name != course.full_name:
                        flash(_('<strong>Richtige Notenliste ausgewählt?</strong><br>Der Kursname "{}" in der Notenliste stimmt nicht mit dem Kurs "{}" überein.'
                                ' Wurde die richtige Notenliste hochgeladen? Bitte überprüfen!'.format(
                            read_course_name, course.full_name)), 'warning')

        # grade import
        rawdata_sheet = grade_wb['RAWDATA']
        # in rawdata mails are saved in column H, starting from line 2
        mail_col = 'H'
        grade_col = app.config['DEFAULT_GRADE_COLUMN']
        max_row = 36  # ToDo: define in config

        if course.language.import_format_id is not None:
            import_format = models.ImportFormat.query.get(course.language.import_format_id)
            if import_format:
                grade_col = import_format.grade_column

        # convert column letters to integer
        mail_col_idx = column_index_from_string(mail_col)
        grade_col_idx = column_index_from_string(grade_col)

        # warnings importing the grades, tuple: (work sheet, coordinate, text)
        # rawdata -> 0, Notenliste -> 1
        warnings = []

        success = 0

        # simultaneously iterate over mail column in RAWDATA sheet and grade column in Notenliste sheet
        for mail_row, grade_row in zip(
            rawdata_sheet.iter_rows(
                min_col=mail_col_idx,
                max_col=mail_col_idx,
                min_row=2,
                max_row=max_row),
            grade_sheet.iter_rows(
                min_col=grade_col_idx,
                max_col=grade_col_idx,
                min_row=2,
                max_row=max_row)
        ):
            # rows returned as a tuple, obtain first and only element
            read_mail = mail_row[0].value
            read_grade = grade_row[0].value

            if read_mail is None or read_grade is None:
                if read_grade is None:
                    warnings.append((1, grade_row[0].coordinate, _('Note fehlt für Kursteilnehmer mit der E-Mail: "{}"'.format(read_mail))))
                continue

            read_mail = str(read_mail).strip().lower()

            # check for valid mail and grade
            if not validate_email(read_mail):
                warnings.append((0, mail_row[0].coordinate, _('Ungültige E-Mail-Adresse: "{}"'.format(read_mail)))
                                )
            elif not is_valid_float(read_grade) or to_float(read_grade) < 0 or to_float(read_grade) > 100:
                warnings.append((1, grade_row[0].coordinate, _('Ungültige Note (keine Zahl zwischen 0 und 100): "{}"'.format(read_grade)))
                                )
            else:
                # add grade to course participant if both - mail and grade - are present in xls file
                applicant = models.Applicant.query.filter_by(mail=read_mail).first()
                if applicant:
                    attendance = course.get_course_attendance(course.id, applicant.id)
                    if attendance:
                        attendance.grade = to_float(read_grade)
                        success += 1
                    else:
                        warnings.append((1, grade_row[0].coordinate,
                                         _('{} ist kein aktiver Kursteilnehmer von {}.'.format(applicant.full_name, course.full_name))))
                else:
                    warnings.append((0, mail_row[0].coordinate,
                                     _('Kursteilnehmer(in) mit folgender Email "{}" wurde nicht gefunden.'.format(
                                         read_mail))))

        # flash warnings
        warning_str = "<strong>Warnungen</strong><br>"
        for warning in warnings:
            sheet = 'RAWDATA' if warning[0] == 0 else 'Notenliste'
            warning_str += f'Blatt "{sheet}" - {warning[1]}: {warning[2]}<br>'
        if len(warnings) > 0:
            flash(warning_str, "warning")

        grade_wb.close()

        return success
